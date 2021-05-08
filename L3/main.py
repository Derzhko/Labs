from math import cos, pi, fabs, sin, radians
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import xlrd

time = 0.1

def create_graph(x, y, name_graph, title_x, title_y):
    fig = plt.figure()
    plt.title(f"{name_graph}")  # заголовок
    plt.xlabel(f"{title_x}")  # ось абсцисс
    plt.ylabel(f"{title_y}")  # ось ординат
    plt.grid()  # включение отображение сетки
    plt.plot(x, y)  # построение графика
    fig.savefig(f'{name_graph}.png')


# Give the location of the file
new_array_L_0 = []
new_array_H_0 = []
new_array_v_x = []
new_array_v_y = []
new_array_v = []
new_array_a = []
array_time = [0]
my_path = "Graph.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_row = my_sheet_obj.max_row
for i in range(2, my_row + 1):
    new_array_L_0.append(my_sheet_obj.cell(row = i, column = 1).value)
for i in range(2, my_row + 1):
    new_array_H_0.append(my_sheet_obj.cell(row = i, column = 2).value)
for i in range(2, my_row + 1):
    new_array_v_x.append(my_sheet_obj.cell(row = i, column = 3).value)
for i in range(2, my_row + 1):
    new_array_v_y.append(my_sheet_obj.cell(row = i, column = 4).value)
for i in range(2, my_row + 1):
    new_array_a.append(my_sheet_obj.cell(row = i, column = 5).value)
for i in range(len(new_array_v_x)):
    new_array_v.append(pow(pow(new_array_v_x[i], 2) + pow(new_array_v_y[i], 2), 0.5))
for i in range(1, len(new_array_v)):
    array_time.append(array_time[i - 1] + time)
# Построение графика


create_graph(new_array_L_0, new_array_H_0, 'траектория', 'L', 'H')
create_graph(array_time, new_array_v, "График скорости от времени", 't', 'v')
create_graph(array_time, new_array_a, "График ускорения от времени", 't', 'a')


#array_H_0 = pd.read_excel("./Graph.xlsx", usecols=1)
#array_H_0.head()