from pulp import *
import openpyxl
variant = 4
count_of_product = 6

profit = []

available = []

collect = []

x1 = pulp.LpVariable("Морковь", lowBound=0)
x2 = pulp.LpVariable("Капуста", lowBound=0)
x3 = pulp.LpVariable("Картофель", lowBound=0)
x4 = pulp.LpVariable("Горох", lowBound=0)
x5 = pulp.LpVariable("Продукт М", lowBound=0)
x6 = pulp.LpVariable("А орэндж", lowBound=0)

my_path = "FARMER.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_column = my_sheet_obj.max_column
row = 2 + variant
for i in range(2, my_column + 1):
    if i < 2 + count_of_product:
        collect.append(float(my_sheet_obj.cell(row=row, column=i).value))
    elif 2 + count_of_product <= i < 2 + 2 * count_of_product:
        available.append(float(my_sheet_obj.cell(row=row, column=i).value))
    else:
        profit.append(float(my_sheet_obj.cell(row=row, column=i).value))

problem = pulp.LpProblem('0',LpMaximize)
problem += collect[0] * profit[0] * x1 + collect[1] * profit[1] * x2 + collect[2] * profit[2] * x3 + collect[3] * profit[3] * x4 + collect[4] * profit[4] * x5 + collect[5] * profit[5] * x6, "Функция цели"
problem += x1 <= available[0], "1"
problem += x2 <= available[1], "2"
problem += x3 <= available[2], "3"
problem += x4 <= available[3], "4"
problem += x5 <= available[4], "5"
problem += x6 <= available[5], "6"
problem += x1 + x2 + x3 + x4 + x5 + x6 <= 10, "Ограничение по посадке"

problem.solve()
print("Результат:")
for variable in problem.variables():
    print(f"{variable.name} = {variable.varValue}")
print(f"Прибыль: {value(problem.objective)}")